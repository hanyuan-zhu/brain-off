"""
导出工具 - 将清单导出为Excel格式

支持标准工程量清单格式
"""

from typing import Dict, Any, Optional
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from src.infrastructure.database.connection import get_session
from models import AnalysisPlan, BOQItem


def export_boq_to_excel(
    plan_id: str,
    output_path: Optional[str] = None
) -> Dict[str, Any]:
    """
    导出清单到Excel文件

    Args:
        plan_id: 计划ID
        output_path: 输出文件路径（可选，默认自动生成）

    Returns:
        Dict包含：
        - success: bool
        - data: {file_path, item_count, total_price}
        - error: str
    """
    try:
        session = get_session()

        # 获取计划信息
        plan = session.query(AnalysisPlan).filter_by(id=plan_id).first()
        if not plan:
            return {
                "success": False,
                "error": f"计划不存在: {plan_id}"
            }

        # 获取清单项
        items = session.query(BOQItem).filter_by(plan_id=plan_id).all()

        if not items:
            return {
                "success": False,
                "error": "清单为空，无法导出"
            }

        # 生成输出路径
        if not output_path:
            export_dir = os.getenv("EXPORT_DIR", "./exports")
            os.makedirs(export_dir, exist_ok=True)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = os.path.join(export_dir, f"BOQ_{plan.project_name}_{timestamp}.xlsx")

        # 创建Excel工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "工程量清单"

        # 设置列宽
        ws.column_dimensions['A'].width = 15  # 编码
        ws.column_dimensions['B'].width = 40  # 名称
        ws.column_dimensions['C'].width = 10  # 单位
        ws.column_dimensions['D'].width = 15  # 工程量
        ws.column_dimensions['E'].width = 15  # 单价
        ws.column_dimensions['F'].width = 15  # 合价

        # 标题行样式
        header_font = Font(name='宋体', size=12, bold=True)
        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        header_alignment = Alignment(horizontal='center', vertical='center')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # 写入标题
        headers = ['项目编码', '项目名称', '单位', '工程量', '单价(元)', '合价(元)']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border

        # 写入数据
        total_price = 0.0
        for row, item in enumerate(items, start=2):
            ws.cell(row=row, column=1, value=item.code or '')
            ws.cell(row=row, column=2, value=item.name)
            ws.cell(row=row, column=3, value=item.unit)
            ws.cell(row=row, column=4, value=item.quantity)
            ws.cell(row=row, column=5, value=item.unit_price)
            ws.cell(row=row, column=6, value=item.total_price)

            if item.total_price:
                total_price += item.total_price

            # 应用边框
            for col in range(1, 7):
                ws.cell(row=row, column=col).border = border

        # 添加合计行
        total_row = len(items) + 2
        ws.cell(row=total_row, column=1, value='合计')
        ws.cell(row=total_row, column=6, value=round(total_price, 2))

        for col in range(1, 7):
            cell = ws.cell(row=total_row, column=col)
            cell.font = Font(bold=True)
            cell.border = border

        # 保存文件
        wb.save(output_path)

        return {
            "success": True,
            "data": {
                "file_path": output_path,
                "item_count": len(items),
                "total_price": round(total_price, 2)
            }
        }

    except Exception as e:
        return {
            "success": False,
            "error": f"导出Excel失败: {str(e)}"
        }


# 导出工具定义
EXPORT_TOOL_DEFINITIONS = [
    {
        "name": "export_boq_to_excel",
        "description": "将工程量清单导出为Excel文件",
        "input_schema": {
            "type": "object",
            "properties": {
                "plan_id": {
                    "type": "string",
                    "description": "计划ID"
                },
                "output_path": {
                    "type": "string",
                    "description": "输出文件路径（可选）"
                }
            },
            "required": ["plan_id"]
        }
    }
]
